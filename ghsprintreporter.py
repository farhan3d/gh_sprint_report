# release 1.2

from __future__ import division
from github3 import login
import datetime
from string import letters
from openpyxl import Workbook
import time
import threading
import smtplib
import csv
from email.mime.text import MIMEText
import app_ui


# the csv file is formatted as follows:
# github username | user email | user manager's username | team name
# this 3 column csv is used to get the emails of the
# users who have violated the commit message format, and
# an email is sent to them and their manager about the
# violating commit. the csv file should reside in the
# same folder this program is being run from.
CSV_FILE_NAME = 'team'
NUM_BUSINESS_DAYS_PER_WEEK = 5


# a class for hourly burndown data structure and xls publishing
class Burndown:

    # process the ideal hours array incrementally and push it to a date-hours dict
    def process_ideal_by_inc(self, estimate_inc):
        self.estimate += estimate_inc
        self.interval = round(float(self.estimate / self.days), 3)
        temp = self.interval
        inc_date = self.start_date
        for num in range(self.days):
            self.date_hours_ideal_map[inc_date] += estimate_inc
            self.date_hours_ideal_map[inc_date] = self.estimate - temp
            temp += self.interval
            if inc_date.weekday() == 4:
                inc_date += datetime.timedelta(days=3)
            else:
                inc_date += datetime.timedelta(days=1)

    def process_actual_item(self, actual_hours, date):
        if (date.weekday() != 5) and (date.weekday() != 6):
            self.date_hours_actual_map[date] += actual_hours
            self.date_hours_burnup_map[date] += actual_hours

    def post_process(self):
        inc_date = self.start_date
        prev_date = None
        prev_hours = self.estimate
        burnup_prev_hours = 0
        for num in range(self.days):
            if (inc_date.weekday() != 5) and (inc_date.weekday() != 6):
                if prev_date:
                    prev_hours = self.date_hours_actual_map[prev_date]
                    burnup_prev_hours = self.date_hours_burnup_map[prev_date]
                temp = prev_hours - self.date_hours_actual_map[inc_date]
                self.date_hours_actual_map[inc_date] = temp
                self.date_hours_burnup_map[inc_date] += burnup_prev_hours
                prev_date = inc_date
            if inc_date.weekday() == 4:
                inc_date += datetime.timedelta(days=3)
            else:
                inc_date += datetime.timedelta(days=1)

    def burndown_data_to_sheet_obj(self, sheet):
        arr = ["Ideal", "Burndown", "Burnup"]
        sheet.add_data_row_bd(arr)
        inc_date = self.start_date
        for num in range(self.days):
            arr = [self.date_hours_ideal_map[inc_date],
                   self.date_hours_actual_map[inc_date],
                   self.date_hours_burnup_map[inc_date]]
            sheet.add_data_row_bd(arr)
            if inc_date.weekday() == 4:
                inc_date += datetime.timedelta(days=3)
            else:
                inc_date += datetime.timedelta(days=1)

    def print_completed_burndown(self):
        inc_date = self.start_date
        for num in range(self.days):
            if inc_date.weekday() == 4:
                inc_date += datetime.timedelta(days=3)
            else:
                inc_date += datetime.timedelta(days=1)

    def __init__(self, start_date, end_date):
        self.start_date = start_date
        self.end_date = end_date
        # this dict will map date keys to an array of ideal and actual hours
        # data structure ---> { date : [ideal_hours, actual_hours] }
        self.date_hours_ideal_map = {}
        self.date_hours_actual_map = {}
        self.date_hours_burnup_map = {}
        self.days = 1
        self.start_date = start_date
        self.end_date = end_date
        self.curr_actual_remaining = 0
        self.estimate = 0
        self.interval = 0
        temp_date = self.start_date
        if self.start_date < self.end_date:
            while temp_date != self.end_date:
                if (temp_date.weekday() != 5) and (temp_date.weekday() != 6):
                    self.days += 1
                temp_date += datetime.timedelta(days=1)
        inc_date = self.start_date
        for num in range(self.days):
            # self.date_hours_map_dict[inc_date] = [0, 0]
            self.date_hours_ideal_map[inc_date] = 0
            self.date_hours_actual_map[inc_date] = 0
            self.date_hours_burnup_map[inc_date] = 0
            if inc_date.weekday() == 4:
                inc_date += datetime.timedelta(days=3)
            else:
                inc_date += datetime.timedelta(days=1)


class ReportSheet:
    def __init__(self, name):
        self.name = name
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = 'Sprint Report'
        self.data = []
        self.bd_data = []
        # if (teamArr):
        #     for item in teamArr:
        #         self.bd_ws = self.wb.create_sheet(str(item) + ' ' + 'Burndown')
        # else:
        self.bd_ws = self.wb.create_sheet('Burndown')
        self.wb.save(name + ".xlsx")

    def add_data_row(self, arr):
        self.data.append(arr)

    def add_data_row_bd(self, arr):
        self.bd_data.append(arr)

    def post_process(self):
        for item in self.data:
            self.ws.append(item)
        for item in self.bd_data:
            self.bd_ws.append(item)
        self.wb.save(self.name + ".xlsx")


# email notification on report generation completion as traversing the
# issues in the repository can take time
def push_email(ui_obj):
    try:
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.ehlo()
        server.starttls()
        server.login(ui_obj.email_input.get(), ui_obj.email_pwd_input.get())
        from_txt = ui_obj.email_input.get()
        to_txt = ui_obj.recipent_input.get()
        subject_txt = "NOTIFICATION: Sprint Report Generated"
        txt = "Hello, your sprint report has been generated. Enjoy!"
        message = """From: %s\nTo: %s\nSubject: %s\n\n%s
        """ % (from_txt, ", ".join(to_txt), subject_txt, txt)
        server.sendmail(ui_obj.email_input.get(), ui_obj.recipent_input.get(), message)
        server.quit()
    except Exception:
        app_ui.update_status_message("Unable to send email", ui_obj ,2)


def push_email_to_user(ui_obj, sender_email, sender_pwd, recipent_email_list, email_sub,
                       email_msg, bcc_email=None, error_code=2):
    try:
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.ehlo()
        server.starttls()
        server.login(sender_email, sender_pwd)
        recipents = list([])
        recipents.extend(recipent_email_list)
        if bcc_email is not None:
            recipents.append(bcc_email)
        message = MIMEText(email_msg)
        message['Subject'] = email_sub
        message['From'] = sender_email
        message['To'] = ", ".join(recipents)
        server.sendmail(sender_email, recipents, message.as_string())
        server.quit()
    except Exception:
        app_ui.update_status_message("Unable to send email", ui_obj, error_code)


# verify that the milestone assigned to the issue is in the on-going sprint
# this will be done by getting the current sprint, extracting the date from
# the current sprint, and comparing against milestone assigned to the issue
def verify_milestone(repo):
    sprint_due_date = None
    open_stones = repo.milestones('open')
    for stone in open_stones:
        if 'Sprint' in str(stone):
            if stone.due_on:
                sprint_due_date = stone.due_on.date()
    curr_date = datetime.date.today()
    if curr_date <= sprint_due_date:
        is_within_sprint = True
    else:
        is_within_sprint = False
    return is_within_sprint


def get_repo_by_index(gh, index):
    repos = gh.repositories()
    reparr = list([])
    for repo in repos:
        reparr.append(repo)
    return reparr[index]


def get_repo_by_name(ui_obj, gh):
    ret_repo = None
    try:
        repos = gh.repositories()
        ret_repo = None
        for repo in repos:
            if repo.name == ui_obj.repo_input.get():
                ret_repo = repo
    except Exception:
        app_ui.update_status_message("Incorrect username / password / repo", ui_obj, 2)
    return ret_repo


def get_sprint_from_issue(issue):
    curr_sprint = issue.milestone
    sprint_info = None
    if curr_sprint:
        sprint_end_date = curr_sprint.due_on.date()
        start_date = sprint_end_date
        sprint_issues_count = curr_sprint.open_issues_count \
            + curr_sprint.closed_issues_count
        sprint_info = {'object': curr_sprint, 'issue-count': sprint_issues_count,
                       'end-date': sprint_end_date, 'start-date': start_date}
    return sprint_info


# this will return a dictionary containing the actual sprint object,
# the number of issues within the sprint, and the due date.
def get_curr_sprint_info(ui_obj, repo):
    curr_sprint = None
    sprint_info = None
    open_stones = repo.milestones('open')
    closed_stones = repo.milestones('closed')
    criteria = 'Sprint'
    num_days = int(ui_obj.sprint_weeks_input.get()) * NUM_BUSINESS_DAYS_PER_WEEK
    if ui_obj.sprint_override_input.get() != '':
        criteria = ui_obj.sprint_override_input.get()
    for stone in closed_stones:
        if criteria in str(stone):
            curr_sprint = stone
    for stone in open_stones:
        if criteria in str(stone):
            curr_sprint = stone
    if curr_sprint:
        sprint_issues_count = curr_sprint.open_issues_count \
            + curr_sprint.closed_issues_count
        sprint_end_date = curr_sprint.due_on.date()
        start_date = sprint_end_date
        for num in range(num_days - 1):
            if start_date.weekday() == 0:
                start_date += datetime.timedelta(days=-3)
            else:
                start_date += datetime.timedelta(days=-1)
        sprint_info = {'object': curr_sprint, 'issue-count': sprint_issues_count,
                       'end-date': sprint_end_date, 'start-date': start_date}
    return sprint_info


def is_date_within_sprint(sprint_info, verify_date):
    if (verify_date <= sprint_info['end-date']) and \
            (verify_date >= sprint_info['start-date']):
        return True
    else:
        return False


def is_date_within_range(start_date, end_date, verify_date):
    if (verify_date >= start_date) and (verify_date <= end_date):
        return True
    else:
        return False


def get_date_from_input(date_str):
    date_year = int(date_str[0:4])
    date_month = int(date_str[5:7])
    date_day = int(date_str[8:10])
    created_date = datetime.date(date_year, date_month, date_day)
    if type(created_date) == datetime.date:
        return created_date
    else:
        return None


# returns the github username of the person who made the comment
# on the sprint issue
def get_comment_author(comment):
    author = comment.user
    return author


# hours can be written as a part of the comment body as
# '8hrs'. the numeric value will be extracted and recorded
# for the resource against the issue, along with additional
# development notes provided inside the issue. how cool is that?
# multiple hour entries in one comment or issue body will be
# added up and written as one value in a row on the sheet.
def parse_comment(text):
    data = [0, None]
    loc = list([])
    hours_text_arr = []
    splt = text.split(' ')
    breakout = False
    for item in splt:
        if breakout is False:
            if 'hrs' in item:
                splt_yet_again = item.split("\n")
                for sub_item in splt_yet_again:
                    if 'hrs' in sub_item:
                        loc.append(sub_item.find('hrs'))
                        full_hours_text = str(sub_item)
                        hours_text_arr.append(full_hours_text)
    for item in hours_text_arr:
        if len(loc) > 0:
            prefix_cleaned = item[:-3]
            prefix_cleaned = prefix_cleaned.translate(None, letters)
            if prefix_cleaned.isdigit():
                data[0] += int(prefix_cleaned)
                data[1] = text.split(item)[1]
    return data


# write information passed through the array into the xlsx
def process_sheet(ws, wb, arr, sheet_data_arr):
    ws.append(arr)
    sheet_data_arr.append(arr)
    wb.save('report.xlsx')
    return sheet_data_arr


def is_item_in_sheet(sheet_data_arr, item, col_num):
    found = False
    if sheet_data_arr:
        for row in sheet_data_arr:
            if row[col_num] == item:
                found = True
                break
    return found


# function checks whether an issue within the sprint or the date range
# has comments or not, then parses any available comments to check for
# hours which are appended to the worksheet array as well as the sheet
# itself.
def process_comments_and_report(sheet, issue, comments, issue_retrieval_method_var,
                                sprint_info=None, start_date=None, end_date=None, bd=None):
    processed_count = 0
    cmnt_count = 0
    some_hours_found = False
    stry_pnts = get_sp(issue)
    assignees = get_assignee_str(issue)
    status = issue.state
    sprint = str(issue.milestone)
    est = get_issue_estimate(issue)
    if bd:
        bd.process_ideal_by_inc(est)
    for comment in comments:
        cmnt_count += 1
        comment_body = comment.body
        hours, notes = parse_comment(comment_body)
        if (hours is not None) and (hours != 0):
            some_hours_found = True
    if (cmnt_count > 0) and some_hours_found:
        for comment in comments:
            comment_date = comment.created_at.date()
            process = False
            if sprint_info and (issue_retrieval_method_var.get() == 1):
                if is_date_within_sprint(sprint_info, comment_date):
                    process = True
            elif issue_retrieval_method_var.get() == 2:
                if start_date and end_date:
                    if is_date_within_range(start_date, end_date, comment_date):
                        process = True
            if process:
                comment_body = comment.body
                comment_id = comment.id
                hours, notes = parse_comment(comment_body)
                if (hours is not None) and (hours != 0):
                    if not is_item_in_sheet(sheet.data, comment_id, 1):
                        arr = [issue.number, assignees, status, stry_pnts,
                               comment_id, str(comment.user), sprint, est, hours,
                               comment_date, notes]
                        sheet.add_data_row(arr)
                        processed_count += 1
                        if bd:
                            bd.process_actual_item(hours, comment_date)
    else:
        # put the issue in the list anyway even if it doesn't have any comments
        # only for the case when a sprint report is required, and not a date
        # range report. For the case of a date range report, this would not be
        # applicable since we need only report issues which have had comments
        # and they can be issues with or without sprint assignments.
        if sprint_info and (issue_retrieval_method_var.get() == 1):
            if not is_item_in_sheet(sheet.data, issue.number, 0):
                arr = [issue.number, assignees, status, stry_pnts, None, None, sprint,
                       est, None, None, None]
                sheet.add_data_row(arr)
                processed_count += 1
    if processed_count > 0:
        return True
    else:
        return False


def get_team_dict_from_csv():
    team_dict = {}
    team_file = open(CSV_FILE_NAME + '.csv', 'rU')
    reader = csv.reader(team_file, dialect=csv.excel_tab)
    for item in reader:
        splt = item[0].split(",")
        team_dict[splt[0]] = [splt[1], splt[2]]
    return team_dict


# checks we need to consider in commit:
# 1) check if issue reference format is present in the commit
# 2) check if 'what, why, impact' are present in the commit message
def is_commit_format(cmt, issue_criteria):
    violation_code = 0
    if (not str(issue_criteria.get()).lower() in str(cmt).lower()) \
        and (not str("Merge").lower() in str(cmt).lower()) \
            and (not str("Rebasing").lower() in str(cmt).lower()):
        violation_code = 1
    return violation_code


# this function will return the story points that are assigned to the
# issue passed to the function. The assignment is done in the form of
# a label on Github in the format '1sp', or '5sp', etc.
def get_sp(issue):
    labels = issue.labels()
    sp_num = 0
    for label in labels:
        if 'sp' in label.name:
            sp_num = int(label.name[:-2])
    return sp_num


def get_assignee_str(issue):
    assignees = issue.assignees
    asg_arr = ''
    count = 0
    for assignee in assignees:
        if count > 0:
            asg_arr += ", "
        asg_arr += str(assignee)
        count += 1
    return str(asg_arr)


# this method assumes that the estimate is provided in the main description
# in the format '4hrs', and that only one estimate is available. If multiple
# estimates are available in the main description, all of the estimates will
# be added up to form the total estimate for the issue / story
def get_issue_estimate(issue):
    hours, notes = parse_comment(issue.body)
    return hours


def gh_login(ui_obj):
    me = login(str(ui_obj.username_input.get()), str(ui_obj.password_input.get()))
    return me


def commits_email_content(repo, cmt):
    email_sub = "[commit message violation] " + \
                str(repo.name) + " " + str(cmt.sha)
    email_msg = "Dear " + str(cmt.author) + \
                ", you have not included the reference to the Github" + \
                " issue in the prescribed format."
    email_msg += "\n\nPlease visit the link below and add the" + \
                 " issue reference in the comment box."
    email_msg += "\n\nCommit URL: " + str(cmt.html_url)
    return email_sub, email_msg


def commits_report(ui_obj):
    gh = gh_login(ui_obj)

    def process_commmit_thrd():
        repo = get_repo_by_name(ui_obj, gh)
        cmt_date = "2018-01-01"
        if ui_obj.commits_date_input.get():
            cmt_date = ui_obj.commits_date_input.get()
        cmts = repo.commits(None, None, None, -1, None, cmt_date, None, None)
        team_dict = get_team_dict_from_csv()
        if team_dict:
            for cmt in cmts:
                msg = 'Processing, please wait...'
                app_ui.update_status_message(msg, ui_obj, 4)
                violation_code = is_commit_format(cmt.commit.message, ui_obj.issue_criteria_input)
                if violation_code == 1:
                    if str(cmt.author) in team_dict:
                        author_email = team_dict[str(cmt.author)][0]
                        author_manager = team_dict[str(cmt.author)][1]
                        manager_email = team_dict[author_manager][0]
                        emails_list = list([])
                        emails_list.append(author_email)
                        if manager_email:
                            emails_list.append(manager_email)
                        comments = cmt.comments()
                        comment_adjustment_found = False
                        email_sub, email_msg = commits_email_content(repo, cmt)
                        for cmnt in comments:
                            if (str(ui_obj.issue_criteria_input.get()).lower()
                                    in str(cmnt.body).lower()):
                                comment_adjustment_found = True
                        if comment_adjustment_found is False:
                            push_email_to_user(ui_obj.commits_sender_email_input.get(),
                                               ui_obj.commits_sender_pwd_input.get(), emails_list,
                                               email_sub, email_msg,
                                               ui_obj.commits_admin_email_input.get(), 6)
                            time.sleep(5.0)
                elif violation_code == 2:   # something for later
                    pass
                elif violation_code == 3:   # something for later
                    pass
                elif violation_code == 4:   # something for later
                    pass
            app_ui.update_status_message("Commit messages processed!", ui_obj, 5)
        else:
            app_ui.update_status_message("Unable to process team CSV!", ui_obj, 2)
    t = threading.Thread(target=process_commmit_thrd)
    t.start()


def team_check(issue, team_name):
    team_name = str(team_name).lower()
    labels = issue.labels()
    is_team_issue = False
    for lbl in labels:
        if team_name == str(lbl.name).lower():
            is_team_issue = True
    return is_team_issue


def sprint_report_main(ui_obj):
    if ui_obj.issue_retrieval_method_var.get() == 1:
        if (ui_obj.sprint_override_input.get() != '') and \
                (ui_obj.sprint_weeks_input.get() != '') and \
                (int(ui_obj.sprint_weeks_input.get()) > 0):
            sprint_report(ui_obj)
        else:
            app_ui.update_status_message("Enter sprint name and weeks count", ui_obj, 2)
    elif ui_obj.issue_retrieval_method_var.get() == 2:
        if (ui_obj.start_date_input.get() != '') and (ui_obj.end_date_input.get() != ''):
            sprint_report(ui_obj)
        else:
            app_ui.update_status_message("Enter start and end dates", ui_obj, 2)


def sprint_report_preprocess(ui_obj, repo):
    repo_issues = repo.issues(None, 'all')
    sprint_info = None
    bd = None
    terminate = False
    if ui_obj.issue_retrieval_method_var.get() == 1:
        sprint_info = get_curr_sprint_info(ui_obj, repo)
        if sprint_info is None:
            app_ui.update_status_message("Invalid sprint name!", ui_obj, 2)
            terminate = True
        else:
            bd = Burndown(sprint_info['start-date'], sprint_info['end-date'])
    else:
        if (get_date_from_input(ui_obj.start_date_input.get())) and \
                get_date_from_input(ui_obj.end_date_input.get()):
            bd = Burndown(get_date_from_input(ui_obj.start_date_input.get()),
                          get_date_from_input(ui_obj.end_date_input.get()))
        else:
            app_ui.update_status_message("Invalid start or end dates!", ui_obj, 2)
            terminate = True
    return repo_issues, sprint_info, bd, terminate


def sprint_report_issue_processor(ui_obj, issue, sprint_info, issues_count_inc,
                                  report_sheet, bd, processed_count):
    msg = 'Processing #' + str(issue.number) + ', please wait...'
    break_out = False
    app_ui.update_status_message(msg, ui_obj, 1)
    if ui_obj.issue_retrieval_method_var.get() == 1:
        if issue.milestone:
            if sprint_info:
                if issue.milestone == sprint_info['object']:
                    issues_count_inc += 1
                    comments = issue.comments()
                    is_processed = process_comments_and_report(report_sheet,
                                                               issue,
                                                               comments,
                                                               ui_obj.issue_retrieval_method_var,
                                                               sprint_info, None,
                                                               None, bd)
                    if is_processed:
                        processed_count += 1
                    if str(ui_obj.isscount_override_input.get()) != '':
                        print(issues_count_inc)
                        if issues_count_inc == int(ui_obj.isscount_override_input.get()):
                            break_out = True
                    else:
                        if issues_count_inc == sprint_info['issue-count']:
                            break_out = True
    elif ui_obj.issue_retrieval_method_var.get() == 2:
        if (ui_obj.start_date_input.get()) and (ui_obj.end_date_input.get()):
            created_start_date = get_date_from_input(ui_obj.start_date_input.get())
            created_end_date = get_date_from_input(ui_obj.end_date_input.get())
            if created_start_date and created_end_date:
                comments = issue.comments()
                if comments:
                    sprint_info = get_sprint_from_issue(issue)
                    is_processed = process_comments_and_report(report_sheet, issue,
                                                               comments,
                                                               ui_obj.issue_retrieval_method_var,
                                                               sprint_info,
                                                               created_start_date,
                                                               created_end_date, bd)
                    if is_processed:
                        processed_count += 1
            else:
                app_ui.update_status_message("Please provide valid dates", ui_obj, 2)
    return break_out, processed_count, issues_count_inc


def sprint_report(ui_obj):
    ui_obj.status_label['text'] = ''
    ui_obj.status_label.configure(foreground="red")
    gh = gh_login(ui_obj)
    report_sheet = ReportSheet('report')
    arr = ["Issue", "Assignees", "Status", "St. Pts", "Comment ID", "Author",
           "Sprint", "Hours Est.", "Actual Hours", "Date", "Comments"]
    report_sheet.add_data_row(arr)

    def process_thread():
        app_ui.update_status_message("Processing, please wait...", ui_obj, 1)
        ui_obj.status_label.configure(foreground="orange")
        repo = get_repo_by_name(ui_obj, gh)
        if repo:
            repo_issues, sprint_info, bd, terminate = sprint_report_preprocess(ui_obj, repo)
            processed_count = 0
            issues_count_inc = 0
            if terminate is False:
                for issue in repo_issues:
                    if ui_obj.issue_term_input.get() != '':
                        if int(ui_obj.issue_term_input.get()) > int(issue.number):
                            break
                    process = True
                    if str(ui_obj.team_input.get()) != '':
                        if team_check(issue, ui_obj.team_input.get()):
                            process = True
                        else:
                            process = False
                    if process:
                        break_out, processed_count, issues_count_inc = \
                            sprint_report_issue_processor(ui_obj, issue, sprint_info,
                                                          issues_count_inc,
                                                          report_sheet,
                                                          bd, processed_count)
                        if break_out:
                            break
            if processed_count > 0:
                bd.post_process()
                bd.burndown_data_to_sheet_obj(report_sheet)
                report_sheet.post_process()
                app_ui.update_status_message("Sprint report generated!", ui_obj, 0)
                push_email(ui_obj)
            else:
                app_ui.update_status_message("Nothing to process, review criteria", ui_obj, 2)

    t = threading.Thread(target=process_thread)
    t.start()


